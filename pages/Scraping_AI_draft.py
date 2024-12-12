import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import io
import logging
from langchain.llms import OpenAI
from langchain.prompts import PromptTemplate
import os

# Configuración de OpenAI
openai_api_key = os.getenv("OPENAI_API_KEY")
llm = OpenAI(temperature=0, api_key=openai_api_key)

# Estilos personalizados
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
    body, p, h1, h2, h3, h4, h5, h6 {
        font-family: 'Poppins', sans-serif !important;
        color: #001978;
    }
    </style>
    """, 
    unsafe_allow_html=True
)

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Cargar archivo SVG
with open("assets/package_blue.svg", "r") as file:
    icon_svg = file.read()

# Título
st.markdown(f"#### {icon_svg} Buscador Semántico en Sitios Web", unsafe_allow_html=True)

# Subida de archivo y configuración
uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])
selected_column = st.text_input("Nombre de la columna de links", "WEBSITE")
consulta_input = st.text_area("Describe lo que quieres encontrar", "Necesito encontrar información referente a: denuncia, denuncias, canal de denuncias, canal ético, compliance, Channel, ethics, complaint, canaldenuncias, canaletico, etico, ético, código de conducta, code of conduct, whistleblower channel, Reporting channel, Whistleblowing channel, canal de ética, ética, Complaints Channel, Sistema Interno de Información, Canal del informante, Canal de información, Canal de comunicación interno, General conditions of sale, buen gobierno")

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}

def make_request(url):
    try:
        response = requests.get(url, headers=headers, timeout=10, verify=False)
        response.raise_for_status()
        return response
    except requests.exceptions.RequestException as e:
        logging.error(f"Error al acceder a {url}: {e}")
        return None

def buscar_con_ia(texto, consulta, url):
    prompt = PromptTemplate(
        input_variables=["texto", "consulta", "url"],
        template = """
            Lee el siguiente contenido: "{texto}". La consulta es: "{consulta}".
            1. Si el contenido menciona o responde a la consulta, resume la información más relevante en un máximo de 1000 caracteres.
            2. Si es relevante, proporciona enlaces útiles relacionados con la consulta, incluyendo el enlace proporcionado: {url}.
            3. Si no se encuentra información relevante, indica: "No se encontró información relevante relacionada con la consulta".
        """
    )
    respuesta = llm(prompt.format(texto=texto, consulta=consulta, url=url))

    if ("no se encontró información relevante" in respuesta.lower()):
        flag = "❌ No se encontró canal de denuncias"
    else:
        flag = "✔️ Canal de denuncias encontrado o contenido relevante"
    
    return respuesta.strip(), flag

def verificar_url(url):
    if not url.startswith("http://") and not url.startswith("https://"):
        return f"https://{url}", f"http://{url}"
    return url, None

def es_url_valida(url):
    if not url or " " in url or not ("." in url):
        return False
    return True

if st.button("Ejecutar búsqueda") and uploaded_file and selected_column and consulta_input:
    try:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active
        website_column_index = None
        
        for cell in sheet[1]:
            if cell.value == selected_column:
                website_column_index = cell.column
                break

        if not website_column_index:
            st.error(f"Columna '{selected_column}' no encontrada.")
            st.stop()

        result_col_index = sheet.max_column + 1
        sheet.cell(row=1, column=result_col_index, value="Resultado")

        for row in range(2, sheet.max_row + 1):
            url = sheet.cell(row=row, column=website_column_index).value
            if not es_url_valida(url):
                sheet.cell(row=row, column=result_col_index, value="URL inválida o vacía")
                continue

            url_https, url_http = verificar_url(url)
            response = make_request(url_https) or make_request(url_http) if url_http else None

            if response:
                soup = BeautifulSoup(response.content, "html.parser")
                text = soup.get_text()
                resultado, flag = buscar_con_ia(text, consulta_input, url_https)
                sheet.cell(row=row, column=result_col_index, value=f"{resultado} - {flag}")
                st.write(f"✔️ Resultado para {url}: {flag}")
            else:
                sheet.cell(row=row, column=result_col_index, value="Error al acceder")
                st.write(f"❌ Error al acceder a {url}")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        st.success("Archivo procesado con éxito.")
        st.download_button(
            label="Descargar archivo procesado",
            data=output,
            file_name="output_with_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
